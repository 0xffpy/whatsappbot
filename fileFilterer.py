import openpyxl
import random


class FileFilterer:
    __listOfNums = ['539118446']
    __file_path = ''
    __listOfNums2 = []
    __numstoremove = ['530498538', '550054434', '504233369', '531050592', '501430811', '580813155', '553226878']

    def __init__(self, number_column=0, filePath=''):
        self.__file_path = filePath
        self.number_column = number_column

    def set_file_path(self, photopath):
        self.__file_path = photopath

    def set_number_of_columns(self, num_columns):
        self.number_column = int(num_columns)

    def excute_program(self):
        self.file = openpyxl.load_workbook(self.__file_path)
        self.fileActive = self.file.active
        self.__gathernumber()
        self.__isreptitive(self.__listOfNums)
        self.__isreptitive(self.__listOfNums2)
        self.__isrep()
        print(self.__listOfNums)
        print(self.__listOfNums2)
        self.__listOfNums = self.__listOfNums + self.__listOfNums2
        self.print_nums(self.__listOfNums)
        self.remove_list()


    def print_nums(self,list):
        for i in list:
            if(i == 0):
                list.remove(i)
                continue
            print(i)
        print(len(list))

    def remove_list(self):
        try:
            for i in self.__numstoremove:
                print('i remove it',i)
                self.__listOfNums.remove(i)
        except:
            return

    def __numfiltter(self, num, index):
        num = str(num)
        num = num.replace(' ', '')
        if '(+966)' in num:
            num = num.lstrip('(+966)')
            num = num[1:]
        if '966' in num:
            num = num.replace('966', '')
        if num[0] == '0':
            num = num[1:]
        if len(num) < 9:
            return 0
        if num[0] == '1' or num[0] == '9':
            return 0
        if len(num) > 9:
            return 0
        return num

    def __gathernumber(self):
        self.maxNumOfRows = self.fileActive.max_row
        for i in range(1, self.maxNumOfRows + 1):
            self.cell = self.fileActive.cell(row=i, column=self.number_column)
            number = self.__numfiltter(self.cell.value, i)
            if number != 0:
                self.__listOfNums.append(number)
            else:
                continue
        for i in range(1, self.maxNumOfRows + 1):
            self.cell = self.fileActive.cell(row=i, column=14)
            number = self.__numfiltter(self.cell.value, i)
            if number != 0:
                self.__listOfNums2.append(number)
            else:
                continue

    def getlistofnums(self):
        return self.__listOfNums


    def __isreptitive(self,list_of_nums):
        for i in range(len(list_of_nums)):
            if list_of_nums[i] == 0:
                continue
            for j in range(i + 1, len(list_of_nums)):
                if list_of_nums[j] == 0:
                    continue
                else:
                    if list_of_nums[i] == list_of_nums[j]:
                        list_of_nums[j] = 0
        for i in list_of_nums:
            if i == 0:
                list_of_nums.remove(i)

    def __isrep(self):
        for i in self.__listOfNums:
            if i in self.__listOfNums2:
                self.__listOfNums.remove(i)

    def generate(self, number_of_lists):
        self.__listOfNums = []
        for i in range(number_of_lists):
            self.__listOfNums.append(self.__generate_number())

    def __generate_number(self):
        return '5' + ['0', '3', '4', '5', '6', '7', '9'][random.randint(0, 6)] + ''.join(
            [str(list(range(0, 10))[random.randint(0, 9)]) if x > -1 else '' for x in list(range(7))])
